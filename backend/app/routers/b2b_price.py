from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from app.routers.b2b_utils import *
from app.auth import require_tier

router = APIRouter()

_WATCH_CATEGORIES = [
    "에어컨", "냉장고", "세탁기", "건조기", "TV",
    "공기청정기", "로봇청소기", "식기세척기", "전기밥솥", "전자레인지",
]



@router.get("/export-report")
async def export_report(
    category: str = Query(..., min_length=1),
    period: str = "3m",
    _: dict = Depends(require_tier("platinum")),
):
    """AI 전략 리포트 데이터를 엑셀(.xlsx)로 다운로드"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _ck = f"{_CACHE_VER}:report:{category}:{period}"
    cached = _GROQ_CACHE.get(_ck)
    report = cached[1] if cached and _time.time() < cached[0] else None

    if not report or not isinstance(report, dict):
        raise HTTPException(status_code=404, detail="리포트 데이터가 없습니다. 먼저 AI 전략 리포트를 생성해주세요.")

    wb = openpyxl.Workbook()

    # ── 스타일 정의 ──────────────────────────────────────────────────────────
    PURPLE = "FF6366F1"
    DARK   = "FF1E1E2E"
    WHITE  = "FFFFFFFF"
    LIGHT  = "FFF8F8FC"
    GREEN  = "FF10B981"

    def _hdr_fill(color=PURPLE):
        return PatternFill("solid", fgColor=color)

    def _thin_border():
        s = Side(style="thin", color="FFD1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _apply_header(ws, row, col, value, bold=True, bg=PURPLE, fg=WHITE, size=11, bg_span=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
        return c

    def _apply_cell(ws, row, col, value, bold=False, color="FF111827", size=10, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=color, size=size)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = _thin_border()
        return c

    today_str = str(date.today())

    # ── Sheet 1: 요약 ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "요약"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 55

    _apply_header(ws1, 1, 1, "항목", bg=PURPLE)
    _apply_header(ws1, 1, 2, "내용", bg=PURPLE)

    rows = [
        ("카테고리", report.get("category", category)),
        ("분석 기간", f"{period} / {today_str}"),
        ("AI 권고 액션", report.get("action", "-")),
        ("권고 근거", report.get("action_reason", "-")),
        ("매입 타이밍", report.get("timing", "-")),
        ("AI 종합 판단", report.get("summary", "-")),
        ("예상 매출 성장", report.get("expected_sales_growth", "-")),
        ("예상 효과 예측", report.get("projection_summary", "-")),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        _apply_cell(ws1, i, 1, k, bold=True, color="FF374151")
        c = ws1.cell(row=i, column=2, value=str(v) if v else "-")
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = _thin_border()
        ws1.row_dimensions[i].height = 36

    # ── Sheet 2: 소비자 분석 ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("소비자 분석")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 60

    _apply_header(ws2, 1, 1, "#", bg=PURPLE)
    _apply_header(ws2, 1, 2, "소비자 니즈", bg=PURPLE)
    for i, item in enumerate(report.get("consumer_needs", []) or [], start=2):
        if item and item not in ("-", "없음"):
            _apply_cell(ws2, i, 1, i - 1, align="center")
            _apply_cell(ws2, i, 2, item)

    ws2.append([""])
    row_off = len(report.get("consumer_needs", [])) + 3
    _apply_header(ws2, row_off, 1, "#", bg="FF059669")
    _apply_header(ws2, row_off, 2, "소비자 불만 키워드", bg="FF059669")
    for i, item in enumerate(report.get("consumer_complaints", []) or [], start=1):
        if item and item not in ("-", "없음"):
            _apply_cell(ws2, row_off + i, 1, i, align="center")
            _apply_cell(ws2, row_off + i, 2, item)

    # ── Sheet 3: 상품 기획 & 기능 제안 ──────────────────────────────────────
    ws3 = wb.create_sheet("상품 기획")
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 60

    _apply_header(ws3, 1, 1, "항목", bg=PURPLE, bg_span=True)
    ws3.merge_cells("A1:B1")
    ws3.cell(1, 1).value = "제품 기획 브리프"

    ws3.cell(2, 1).value = report.get("product_brief", "-")
    ws3.cell(2, 1).font = Font(bold=True, size=11, color="FF6366F1")
    ws3.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws3.merge_cells("A2:B2")
    ws3.row_dimensions[2].height = 48

    ws3.append([""])
    _apply_header(ws3, 4, 1, "#", bg="FF7C3AED")
    _apply_header(ws3, 4, 2, "추천 탑재 기능", bg="FF7C3AED")
    for i, feat in enumerate(report.get("recommended_features", []) or [], start=1):
        if feat and feat not in ("-", "없음"):
            _apply_cell(ws3, 4 + i, 1, i, align="center")
            _apply_cell(ws3, 4 + i, 2, feat)

    # ── Sheet 4: 매입·가격 전략 ───────────────────────────────────────────────
    ws4 = wb.create_sheet("매입·가격 전략")
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 55

    items4 = [
        ("signal", "시장 신호"),
        ("price_strategy", "가격 전략"),
        ("timing", "매입 타이밍"),
        ("strategy", "핵심 전략"),
    ]
    _apply_header(ws4, 1, 1, "구분", bg=PURPLE)
    _apply_header(ws4, 1, 2, "내용", bg=PURPLE)
    for i, (key, label) in enumerate(items4, start=2):
        val = report.get(key) or "-"
        if isinstance(val, list):
            val = "\n".join(f"• {v}" for v in val if v)
        _apply_cell(ws4, i, 1, label, bold=True, color="FF374151")
        c = ws4.cell(row=i, column=2, value=str(val))
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = _thin_border()
        ws4.row_dimensions[i].height = 40

    # ── Sheet 5: 예상 효과 ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("예상 효과")
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 55

    _apply_header(ws5, 1, 1, "#", bg=GREEN)
    _apply_header(ws5, 1, 2, "예상 효과", bg=GREEN)
    for i, eff in enumerate(report.get("expected_effects", []) or [], start=1):
        if eff and eff not in ("-", "없음"):
            _apply_cell(ws5, 1 + i, 1, i, align="center")
            _apply_cell(ws5, 1 + i, 2, eff)

    # ── 바이트 스트림으로 변환 후 반환 ──────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"가전무쌍_B2B리포트_{category}_{today_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )

@router.get("/price-monitor")
async def get_price_monitor(_: dict = Depends(require_b2b)):
    """전체 카테고리 가격 변동 모니터링 — 매입 알림 생성"""
    import json as _json
    from app.database import fetchall

    # ── DB 스냅샷 기반 가격 계산 ──────────────────────────────────────────────
    async def _from_db(category: str):
        rows = await fetchall(
            "SELECT snapshot_date, avg_price, min_price, max_price, brand_data "
            "FROM price_history WHERE category = %s "
            "ORDER BY snapshot_date DESC LIMIT 14",
            (category,),
        )
        if len(rows) < 2:
            return None

        latest   = rows[0]
        prev_day = rows[1]
        week_ago = rows[min(6, len(rows) - 1)]

        # 날짜 차이 확인 — 최소 간격 미달이면 변동폭 신뢰 불가
        from datetime import datetime as _dt2
        def _days_apart(a, b) -> int:
            try:
                da = a["snapshot_date"] if hasattr(a["snapshot_date"], "toordinal") else _dt2.strptime(str(a["snapshot_date"]), "%Y-%m-%d").date()
                db_ = b["snapshot_date"] if hasattr(b["snapshot_date"], "toordinal") else _dt2.strptime(str(b["snapshot_date"]), "%Y-%m-%d").date()
                return abs((da - db_).days)
            except Exception:
                return 0

        _day_gap  = _days_apart(latest, prev_day)
        _week_gap = _days_apart(latest, week_ago)

        # 전일 대비: 1~2일 간격인 경우만 신뢰, 변동폭 ≤8% (일일 가격 변동 현실적 상한)
        _raw_day  = round((latest["avg_price"] - prev_day["avg_price"]) / max(prev_day["avg_price"], 1) * 100, 1)
        day_chg   = _raw_day if (1 <= _day_gap <= 2 and abs(_raw_day) <= 8) else None

        # 7일 대비: 5~10일 간격인 경우만 신뢰, 변동폭 ≤15%
        _raw_week = round((latest["avg_price"] - week_ago["avg_price"]) / max(week_ago["avg_price"], 1) * 100, 1)
        week_chg  = _raw_week if (5 <= _week_gap <= 10 and abs(_raw_week) <= 15) else None

        # ── 브랜드별 변동 ─────────────────────────────────────────────────────
        brand_changes: list = []
        try:
            if latest.get("brand_data") and week_ago.get("brand_data"):
                l_bd = _json.loads(latest["brand_data"])
                w_bd = _json.loads(week_ago["brand_data"])
                # 구 배열 포맷 / 신 dict 포맷 모두 지원
                l_list = l_bd if isinstance(l_bd, list) else l_bd.get("brands", [])
                w_list = w_bd if isinstance(w_bd, list) else w_bd.get("brands", [])
                l_bmap = {b["brand"]: b["avg"] for b in l_list if isinstance(b, dict)}
                w_bmap = {b["brand"]: b["avg"] for b in w_list if isinstance(b, dict)}
                for brand, curr in l_bmap.items():
                    if brand in w_bmap and w_bmap[brand] > 0:
                        chg = round((curr - w_bmap[brand]) / w_bmap[brand] * 100, 1)
                        if abs(chg) >= 2:
                            brand_changes.append({"brand": brand, "change_pct": chg, "current_avg": curr})
                brand_changes.sort(key=lambda x: x["change_pct"])
        except Exception:
            pass

        # ── min_price 정제 ────────────────────────────────────────────────────
        _avg = latest["avg_price"]
        _min = latest["min_price"]
        try:
            if latest.get("brand_data"):
                _bd = _json.loads(latest["brand_data"])
                _bl = _bd if isinstance(_bd, list) else _bd.get("brands", [])
                _b_avgs = [b["avg"] for b in _bl if isinstance(b, dict) and b.get("avg", 0) > 0]
                if _b_avgs:
                    _min = min(_b_avgs)
        except Exception:
            pass
        if _min < _avg * 0.60:
            _min = int(_avg * 0.60)

        return {
            "category":        category,
            "avg_price":       _avg,
            "min_price":       _min,
            "day_change_pct":  day_chg,
            "week_change_pct": week_chg,
            "brand_changes":   brand_changes[:4],
            "last_updated":    str(latest["snapshot_date"]),
            "history": [
                {"date": str(r["snapshot_date"]), "avg_price": r["avg_price"]}
                for r in reversed(rows[:7])
            ],
            "realtime": False,
        }

    # ── 실시간 네이버 쇼핑 조회 (DB 스냅샷 없을 때 폴백) ─────────────────────
    async def _from_realtime(category: str):
        try:
            raw = await search_products(query=category, page=1, display=80, sort="sim", category=category)
            items = raw.get("items", [])
            # search_products가 반환하는 필드는 it["price"]
            prices_raw = [it["price"] for it in items if it.get("price", 0) > 0]
            if not prices_raw:
                return None
            _sorted_p = sorted(prices_raw)
            _median_p = _sorted_p[len(_sorted_p) // 2]
            # 중앙값 50%~300% 범위만 신뢰 (부품·액세서리 제거)
            prices = [p for p in _sorted_p if _median_p * 0.5 <= p <= _median_p * 3.0]
            if not prices:
                prices = _sorted_p
            avg_p = int(sum(prices) / len(prices))
            # 브랜드별 평균가
            from collections import defaultdict as _dd
            brand_prices: dict = _dd(list)
            for it in items:
                if not (_median_p * 0.5 <= it.get("price", 0) <= _median_p * 3.0):
                    continue
                brand = it.get("brand") or it.get("maker") or ""
                if brand:
                    brand_prices[brand].append(it["price"])
            brand_avgs = {b: int(sum(ps) / len(ps)) for b, ps in brand_prices.items() if len(ps) >= 2}
            # min_price = 브랜드 평균가 중 최솟값, 60% 플로어
            if brand_avgs:
                min_p = min(brand_avgs.values())
            else:
                min_p = prices[0]
            if min_p < avg_p * 0.60:
                min_p = int(avg_p * 0.60)

            return {
                "category":        category,
                "avg_price":       avg_p,
                "min_price":       min_p,
                "day_change_pct":  None,
                "week_change_pct": None,
                "brand_changes":   [],
                "last_updated":    str(date.today()),
                "history":         [{"date": str(date.today()), "avg_price": avg_p}],
                "realtime":        True,
            }
        except Exception:
            return None

    # ── 병렬 조회 ─────────────────────────────────────────────────────────────
    db_results = await asyncio.gather(*[_from_db(c) for c in _WATCH_CATEGORIES])

    results = []
    rt_needed = [c for c, r in zip(_WATCH_CATEGORIES, db_results) if r is None]
    rt_results_raw = await asyncio.gather(*[_from_realtime(c) for c in rt_needed])
    rt_map = {c: r for c, r in zip(rt_needed, rt_results_raw) if r}

    for cat, db_r in zip(_WATCH_CATEGORIES, db_results):
        item = db_r or rt_map.get(cat)
        if item is None:
            continue

        week_chg = item.get("week_change_pct")
        if item.get("realtime"):
            signal, signal_type = "실시간 조회", "realtime"
        elif week_chg is None:
            signal, signal_type = "데이터 부족", "neutral"
        elif week_chg <= -5:
            signal, signal_type = "매입 적기",  "buy"
        elif week_chg <= -2:
            signal, signal_type = "하락 추세",  "watch"
        elif week_chg >= 5:
            signal, signal_type = "가격 상승",  "wait"
        elif week_chg >= 2:
            signal, signal_type = "상승 추세",  "caution"
        else:
            signal, signal_type = "보합",       "neutral"

        results.append({**item, "signal": signal, "signal_type": signal_type})

    results.sort(key=lambda x: (x.get("week_change_pct") or 0))

    buy_signals = [r["category"] for r in results if r["signal_type"] == "buy"]
    drop_count  = sum(1 for r in results if r["signal_type"] in ("buy", "watch"))
    rise_count  = sum(1 for r in results if r["signal_type"] in ("wait", "caution"))
    has_realtime = any(r.get("realtime") for r in results)

    return {
        "summary": {
            "total":        len(results),
            "drop_count":   drop_count,
            "rise_count":   rise_count,
            "buy_signals":  buy_signals,
            "has_realtime": has_realtime,
        },
        "categories":  results,
        "updated_at":  str(date.today()),
    }
